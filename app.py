from dotenv import load_dotenv
load_dotenv()
# app.py
import os
import io
import re
import json
from datetime import datetime
from flask import (
    Flask, render_template, request, redirect, url_for, session, flash,
    jsonify, send_file
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, or_
from flask_bcrypt import Bcrypt
import pandas as pd
from flask_admin import Admin

# PDF libs
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.styles import ParagraphStyle
from io import BytesIO

# Excel styling
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------- App config ----------------
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY")   # production!

# Use Neon DB if available, else fall back to SQLite
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:///instance/users.db")
# Force SSL for Neon
if app.config["SQLALCHEMY_DATABASE_URI"].startswith("postgresql"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "connect_args": {"sslmode": "require"}
    }
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

# ---------------- Models ----------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default="user")  # "user" or "admin"

class Tariff(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    dimensions = db.Column(db.String(200))  # optional
    mat_cost = db.Column(db.Float, nullable=False)

class ContainerInfo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    container_no = db.Column(db.String(50))
    line = db.Column(db.String(50))
    size = db.Column(db.String(20))
    in_date = db.Column(db.String(20))
    mfg_date = db.Column(db.String(20))
    gw = db.Column(db.String(20))
    tw = db.Column(db.String(20))
    csc = db.Column(db.String(100))

    reports = db.relationship("Report", backref="container_info", lazy=True)

from sqlalchemy.dialects.postgresql import JSONB

class Report(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False)
    container_no = db.Column(db.String(50), nullable=False)
    line = db.Column(db.String(50), nullable=True)
    file_type = db.Column(db.String(20), nullable=False)
    file_path = db.Column(db.String(300), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    grand_total = db.Column(db.Float, default=0.0)
    container_id = db.Column(db.Integer, db.ForeignKey("container_info.id"))
    
    # 🆕 New field for repair entries
    entries_json = db.Column(JSONB, nullable=True)

with app.app_context():
     db.create_all()

# ---------------- Utilities ----------------
def safe_filename(name: str):
    if not name:
        return "estimation"
    name = str(name).strip()
    # Replace invalid filename characters with underscore
    return re.sub(r'[^A-Za-z0-9_\-\.]', "_", name)

def ensure_user_dir(username):
    base_dir = os.path.join(os.getcwd(), "reports")
    user_dir = os.path.join(base_dir, username)
    os.makedirs(user_dir, exist_ok=True)
    return user_dir

# ---------------- CLI: import tariffs from Excel ----------------
@app.cli.command("import-tariffs")
def import_tariffs():
    """Import Tarrifs_repairs.xlsx into Tariff table (one-time)."""
    excel_path = os.path.join(os.path.dirname(__file__), "Tarrifs_repairs.xlsx")
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found at {excel_path}")
        return

    df = pd.read_excel(excel_path).fillna("")

    # normalize column names
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {"category": None, "description": None, "mat cost": None, "dimensions": None}
    for col in df.columns:
        if "category" in col:
            col_map["category"] = col
        elif "desc" in col:
            col_map["description"] = col
        elif "mat" in col and "cost" in col:
            col_map["mat cost"] = col
        elif "dim" in col:
            col_map["dimensions"] = col

    if not col_map["category"] or not col_map["description"] or not col_map["mat cost"]:
        print("❌ Could not detect required columns (Category, Description, MAT COST).")
        print("Columns found:", df.columns.tolist())
        return

    # clear existing
    Tariff.query.delete()
    db.session.commit()

    for _, row in df.iterrows():
        try:
            mat_val = float(str(row[col_map["mat cost"]]).replace(",", "").strip() or 0)
        except:
            mat_val = 0.0
        t = Tariff(
            category=str(row[col_map["category"]]).strip(),
            description=str(row[col_map["description"]]).strip(),
            dimensions=str(row[col_map["dimensions"]]).strip() if col_map["dimensions"] else "",
            mat_cost=mat_val
        )
        db.session.add(t)
    db.session.commit()

    print("✅ Tariffs imported into DB.")
    print("\n🔎 Preview (first 5 rows):")
    preview = Tariff.query.limit(5).all()
    for p in preview:
        print(f"- {p.category} | {p.description} | {p.dimensions or '(no dim)'} | {p.mat_cost}")

# ---------------- AUTH ----------------
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = User.query.filter_by(username=username).first()
        if not user:
            flash("Invalid username or password.", "error")
            return render_template("login.html")

        if bcrypt.check_password_hash(user.password, password):
            session["user"] = username
            session["user_role"] = user.role
            flash("Login successful!", "success")
            if user.role == "admin":
                return redirect("dashboard")
            else:
                return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "error")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user", None)
    flash("Logged out successfully!", "info")
    return redirect(url_for("login"))

# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))

    username = session["user"]
    user = User.query.filter_by(username=username).first()

       # ✅ Admin Dashboard
    if user.role == "admin":
        # total reports
        total_reports = Report.query.count() or 0

        # Reports generated this month (safely)
        now = datetime.utcnow()
        # month bounds
        start_month = datetime(now.year, now.month, 1)
        if now.month == 12:
            next_month = datetime(now.year + 1, 1, 1)
        else:
            next_month = datetime(now.year, now.month + 1, 1)

        reports_this_month = Report.query.filter(
            Report.timestamp >= start_month,
            Report.timestamp < next_month
        ).count()

        # total unique containers repaired (all time) and for the month
        from sqlalchemy import func
        total_containers = db.session.query(func.count(func.distinct(Report.container_no))).scalar() or 0
        containers_this_month = db.session.query(
            func.count(func.distinct(Report.container_no))
        ).filter(Report.timestamp >= start_month, Report.timestamp < next_month).scalar() or 0

        # Reports by line (all time)
        raw_by_line_all = db.session.query(Report.line, func.count(Report.id)).group_by(Report.line).order_by(func.count(Report.id).desc()).all()
        by_line_all = [[r[0] or "Unknown", int(r[1])] for r in raw_by_line_all] if raw_by_line_all else []

        # Reports by line (this month)
        raw_by_line_month = db.session.query(Report.line, func.count(Report.id)).filter(
            Report.timestamp >= start_month, Report.timestamp < next_month
        ).group_by(Report.line).order_by(func.count(Report.id).desc()).all()
        by_line_month = [[r[0] or "Unknown", int(r[1])] for r in raw_by_line_month] if raw_by_line_month else []

        # Latest 5 reports (newest first)
        latest_reports = Report.query.order_by(Report.timestamp.desc()).limit(5).all()

        # containers by month (last 12 months) -> simple list [label, count]
        months = []
        for i in range(11, -1, -1):
            # compute year/month offset properly
            year = (now.year - ((now.month - 1 - i) // 12))
            month = ((now.month - 1 - i) % 12) + 1
            first_day = datetime(year, month, 1)
            if month == 12:
                next_m = datetime(year + 1, 1, 1)
            else:
                next_m = datetime(year, month + 1, 1)
            cnt = db.session.query(func.count(func.distinct(Report.container_no))).filter(Report.timestamp >= first_day, Report.timestamp < next_m).scalar() or 0
            months.append([first_day.strftime("%b %y"), int(cnt)])
        containers_by_month = months

        # containers by day for current month
        per_day_raw = db.session.query(func.date(Report.timestamp), func.count(func.distinct(Report.container_no))).filter(
            Report.timestamp >= start_month, Report.timestamp < next_month
        ).group_by(func.date(Report.timestamp)).order_by(func.date(Report.timestamp)).all()
        # make map date->count then fill days 1..today.day
        per_day_map = {d.strftime("%Y-%m-%d"): c for d, c in per_day_raw}
        days_list = []
        for d in range(1, now.day + 1):
            cur = datetime(now.year, now.month, d)
            key = cur.strftime("%Y-%m-%d")
            days_list.append([str(d), int(per_day_map.get(key, 0))])
        containers_by_day = days_list

        return render_template(
            "dashboard_admin.html",
            username=username,
            total_reports=total_reports,
            reports_this_month=reports_this_month,
            by_line_all=by_line_all,
            by_line_month=by_line_month,
            total_containers=total_containers,
            containers_this_month=containers_this_month,
            latest_reports=latest_reports,
            containers_by_month=containers_by_month,
            containers_by_day=containers_by_day,
            current_month=start_month.strftime("%Y-%m")
        )

    # ✅ Normal User Dashboard
    else:
        report_count = Report.query.filter_by(username=username).count()
        return render_template("dashboard.html",
                               username=username,
                               report_count=report_count,
                               )

# ---------------- ESTIMATION (PAGE 2) ----------------
# ---------------- ESTIMATION (Page 2) ----------------
@app.route("/estimation", methods=["GET", "POST"])
def estimation():
    if "user" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        container_no = request.form.get("container_no").strip()
        in_date = request.form.get("in_date")
        # ---- Normalize Manufacturing Date (supports YYYY-MM, YYYY-MM-DD) ----
        mfg_date_raw = request.form.get("mfg_date", "").strip()
        mfg_date = ""
        line = request.form.get("line").strip()
        size = request.form.get("size").strip()
        gw = request.form.get("gw")
        tw = request.form.get("tw")
        csc = request.form.get("csc").strip()

        errors = []

        # --- Date validation ---
        try:
            if len(mfg_date_raw) == 7:  # yyyy-mm from <input type="month">
                mfg_date = datetime.strptime(mfg_date_raw, "%Y-%m").strftime("%m/%Y")
            elif len(mfg_date_raw) == 10:  # yyyy-mm-dd from <input type="date">
                mfg_date = datetime.strptime(mfg_date_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
            else:
                mfg_date = mfg_date_raw
        except Exception:
            mfg_date = mfg_date_raw or ""

        # ---- Validate In Date (always yyyy-mm-dd) ----
        in_date_raw = request.form.get("in_date", "").strip()
        errors = []

        try:
            in_date_dt = datetime.strptime(in_date_raw, "%Y-%m-%d")
            # For mfg_date, if only month/year, assume 1st of that month
            if len(mfg_date_raw) == 7:
                mfg_date_dt = datetime.strptime(mfg_date_raw + "-01", "%Y-%m-%d")
            else:
                mfg_date_dt = datetime.strptime(mfg_date_raw, "%Y-%m-%d")
            if mfg_date_dt >= in_date_dt:
                errors.append("Manufacturing Date must be earlier than In Date.")
        except Exception:
            # Only flag invalid In Date, not month/year case
            if len(mfg_date_raw) != 7:
                errors.append("Invalid date format.")

        # --- Weight validation ---
        try:
            gw_val = float(gw)
            tw_val = float(tw)
            if gw_val <= tw_val:
                errors.append("Gross Weight must be greater than Tare Weight.")
        except Exception:
            errors.append("GW and TW must be numeric.")

        # --- Handle errors ---
        if errors:
            for e in errors:
                flash(e, "error")
            return redirect(url_for("estimation"))

        # ✅ Save container info to database
        container = ContainerInfo(
            container_no=container_no,
            line=line,
            size=size,
            in_date=in_date,
            mfg_date=mfg_date,
            gw=gw,
            tw=tw,
            csc=csc
        )
        db.session.add(container)
        db.session.commit()

        # ✅ Save container_id in session for linking reports
        session["container_id"] = container.id

        # ✅ Save all container data in session (for Page 3 prefill)
        session["estimation_data"] = {
            "container_no": container_no,
            "in_date": in_date,
            "mfg_date": mfg_date,
            "line": line,
            "size": size,
            "gw": gw,
            "tw": tw,
            "csc": csc,
        }

        flash("Container details saved successfully!", "success")
        return redirect(url_for("page3"))

    return render_template("estimation.html", username=session["user"])

# ---------------- PAGE 3 ----------------
@app.route("/page3")
def page3():
    if "estimation_data" not in session:
        return redirect(url_for("estimation"))
    categories = [c[0] for c in db.session.query(Tariff.category).distinct().all()]
    categories.sort()
    return render_template("page3.html",
                           username=session["user"],
                           container=session["estimation_data"],
                           categories=categories)

# Add Manual Tariff Entry
@app.route("/add_manual_tariff", methods=["POST"])
def add_manual_tariff():
    if "user" not in session:
        return jsonify({"error": "Not logged in"}), 401

    data = request.get_json() or {}
    category = data.get("category", "").strip()
    description = data.get("description", "").strip()
    dimension = data.get("dimension", "").strip() or None
    mat_cost = data.get("mat_cost", 0)

    if not category or not description:
        return jsonify({"error": "Missing required fields"}), 400

    # Check if same description already exists for this dimension/category
    existing = db.session.execute(
        text("""
            SELECT * FROM tariff
            WHERE Category = :cat AND Description = :desc
              AND COALESCE(Dimensions, '') = COALESCE(:dim, '')
        """),
        {"cat": category, "desc": description, "dim": dimension}
    ).fetchone()

    if existing:
        return jsonify({"message": "Entry already exists in tariff table"}), 200

    # Insert the new tariff entry
    db.session.execute(
        text("""
            INSERT INTO tariff (Category, Description, Dimensions, MAT_COST)
            VALUES (:cat, :desc, :dim, :mat)
        """),
        {"cat": category, "desc": description, "dim": dimension, "mat": mat_cost}
    )
    db.session.commit()

    return jsonify({"message": "Manual repair added to tariff table"}), 200

def get_storage_base():
    """
    Return a base folder for storing exported files.
    Preference order:
      1. environment variable FILE_STORAGE_PATH (configurable)
      2. Render persistent disk mount (/mnt/data) if available
      3. /tmp/reports (writable on Render free tier until restart)
      4. local './reports' folder for dev
    """
    # 1) explicit override (good for production)
    path = os.getenv("FILE_STORAGE_PATH")
    if path:
        return path

    # 2) Render persistent disk (recommended if you've added a disk in Render dashboard)
    if os.path.isdir("/mnt/data"):
        return os.path.join("/mnt/data", "reports")

    # 3) Render free-tier tmp
    if os.path.isdir("/tmp"):
        return os.path.join("/tmp", "reports")

    # 4) fallback to local
    return os.path.join(os.getcwd(), "reports")

def ensure_user_dir(username):
    base = get_storage_base()
    user_dir = os.path.join(base, username)
    os.makedirs(user_dir, exist_ok=True)
    return user_dir

from flask import send_from_directory, abort

@app.route("/preview/<int:report_id>")
def preview_report(report_id):
    from flask import send_from_directory
    from sqlalchemy.orm import joinedload

    # Fetch the report record
    report = Report.query.get_or_404(report_id)

    # Extract file info
    file_path = report.file_path
    if isinstance(file_path, (tuple, list)):
        file_path = file_path[0]

    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)

    # Check if file exists
    if not os.path.exists(file_path):
        app.logger.warning(f"[Auto-regen] File missing for report {report.id}...")

        # Gather related data
        container = ContainerInfo.query.get(report.container_id)
        entries = report.entries_json or []

        # Try to load report details for regeneration
        if report.file_type == "pdf":
            try:
                app.logger.info(f"[Auto-regen] Regenerating missing PDF report {report.id}...")
                new_path, _ = export_pdf_internal(
                    report.username, container, entries, report.line, report.container_no
                )
                report.file_path = new_path
                db.session.commit()
                app.logger.info(f"[Auto-regen] Rebuilt PDF for {report.container_no}")
                file_path = new_path
                directory = os.path.dirname(file_path)
                filename = os.path.basename(file_path)
            except Exception as e:
                app.logger.error(f"[Auto-regen] Failed to regenerate PDF: {e}")
                flash("Could not regenerate missing PDF file.", "error")
                return redirect(url_for("reports"))

        elif report.file_type == "excel":
            try:
                app.logger.info(f"[Auto-regen] Regenerating missing Excel report {report.id}...")
                new_path, _ = export_excel_internal(
                    report.username, container, entries, report.line, report.container_no
                )
                report.file_path = new_path
                db.session.commit()
                app.logger.info(f"[Auto-regen] Rebuilt Excel for {report.container_no}")
                file_path = new_path
                directory = os.path.dirname(file_path)
                filename = os.path.basename(file_path)
            except Exception as e:
                app.logger.error(f"[Auto-regen] Failed to regenerate Excel: {e}")
                flash("Could not regenerate missing Excel file.", "error")
                return redirect(url_for("reports"))

    # Security: ensure file within reports folder
    base = os.path.abspath(get_storage_base())
    file_abspath = os.path.abspath(file_path)
    if not file_abspath.startswith(base):
        abort(403)

    if not os.path.exists(file_abspath):
        flash("File not found or could not be regenerated.", "error")
        return redirect(url_for("reports"))

    app.logger.info(f"[Preview] Serving {filename} from {directory}")
    return send_from_directory(directory, filename)

# ---------------- Reports Page ----------------
@app.route("/reports")
def reports():
    # require login
    if "user" not in session:
        return redirect(url_for("login"))

    username = session["user"]
    user = User.query.filter_by(username=username).first()

    # query params
    q = request.args.get("q", "").strip()
    line = request.args.get("line")
    month = request.args.get("month")    # expected format: "YYYY-MM"
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    # base query
    qs = Report.query

    # non-admins only see their own reports
    if not user or getattr(user, "role", "") != "admin":
        qs = qs.filter_by(username=username)

    # search across container_no and line
    if q:
        qs = qs.filter(
            or_(
                Report.container_no.ilike(f"%{q}%"),
                Report.line.ilike(f"%{q}%"),
            )
        )

    # filter by line
    if line:
        qs = qs.filter(Report.line == line)

    # filter by month (YYYY-MM)
    if month:
        try:
            dt = datetime.strptime(month, "%Y-%m")
            start = datetime(dt.year, dt.month, 1)
            # compute month end (first day of next month)
            if dt.month == 12:
                end = datetime(dt.year + 1, 1, 1)
            else:
                end = datetime(dt.year, dt.month + 1, 1)
            qs = qs.filter(Report.timestamp >= start, Report.timestamp < end)
        except Exception:
            # invalid month format - ignore filter
            pass

    # newest first
    qs = qs.order_by(Report.timestamp.desc())

    # paginate - returns sqlalchemy Pagination object (items, has_next, etc.)
    pagination = qs.paginate(page=page, per_page=per_page, error_out=False)
    reports_list = pagination.items

    # distinct lines for the filter dropdown
    lines_query = db.session.query(ContainerInfo.line).distinct().all()
    lines = [r[0] for r in lines_query if r[0]] if lines_query else []

    return render_template(
        "reports.html",
        reports=reports_list,
        pagination=pagination,
        q=q,
        line=line,
        month=month,
        lines=lines,
        is_admin=(user and getattr(user, "role", "") == "admin"),
    )

@app.route("/delete_report/<int:report_id>", methods=["POST"])
def delete_report(report_id):
    if "user" not in session:
        return jsonify({"error": "Not logged in"}), 401

    username = session["user"]
    user = User.query.filter_by(username=username).first()

    report = Report.query.get(report_id)
    if not report:
        return jsonify({"error": "Report not found"}), 404

    # Only admin or report owner can delete
    if user.role != "admin" and report.username != username:
        return jsonify({"error": "Unauthorized"}), 403

    db.session.delete(report)
    db.session.commit()
    return jsonify({"message": "Report deleted successfully"})

# @app.route("/preview/<int:report_id>")
# def preview_report(report_id):
    report = Report.query.get_or_404(report_id)
    try:
        # Build absolute path safely
        file_path = os.path.abspath(report.file_path)

        # Check file exists
        if not os.path.exists(file_path):
            flash("File not found on server.", "error")
            return redirect(url_for("reports"))

        # Return file inline for browser preview
        if report.file_type == "pdf":
            return send_file(file_path, mimetype="application/pdf")
        elif report.file_type == "excel":
            # Browsers can’t preview Excel → force download instead
            return send_file(file_path, as_attachment=True)
        else:
            flash("Unsupported file type.", "error")
            return redirect(url_for("reports"))

    except Exception as e:
        print("Preview error:", e)
        flash("Could not open file.", "error")
        return redirect(url_for("reports")) 

@app.route("/export_reports_excel")
def export_reports_excel():
    if "user" not in session:
        return redirect(url_for("login"))

    from datetime import datetime
    import pandas as pd
    from io import BytesIO
    from flask import send_file

    username = session["user"]
    user = User.query.filter_by(username=username).first()

    # --- Collect Filters ---
    container_no = request.args.get("container_no", "").strip()
    file_type = request.args.get("file_type", "").strip()
    line_filter = request.args.get("line", "").strip()
    user_filter = request.args.get("user", "").strip()
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # --- Base Query ---
    query = db.session.query(Report, ContainerInfo).join(
        ContainerInfo, Report.container_id == ContainerInfo.id
    )

    if user.role != "admin":
        query = query.filter(Report.username == username)
    elif user_filter:
        query = query.filter(Report.username == user_filter)

    # --- Apply Filters ---
    if container_no:
        query = query.filter(ContainerInfo.container_no.ilike(f"%{container_no}%"))
    if line_filter:
        query = query.filter(ContainerInfo.line.ilike(f"%{line_filter}%"))
    if file_type:
        query = query.filter(Report.file_type.ilike(f"%{file_type}%"))
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            query = query.filter(Report.timestamp.between(start_dt, end_dt))
        except Exception:
            pass

    reports = query.all()
    if not reports:
        flash("No reports match the filters.", "warning")
        return redirect(url_for("reports"))

    # --- Create Excel Data ---
    data = []
    for r, c in reports:
        data.append({
            "Container No": c.container_no,
            "Line": c.line,
            "Size": c.size,
            "In Date": c.in_date,
            "Mfg Date": c.mfg_date,
            "GW": c.gw,
            "TW": c.tw,
            "CSC": c.csc,
            "Grand Total": getattr(r, "grand_total", ""),
            "File Type": r.file_type,
            "Date Generated": r.timestamp.strftime("%Y-%m-%d %H:%M")
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Filtered Reports")
        ws = writer.sheets["Filtered Reports"]
        from openpyxl.styles import Alignment, Font
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    output.seek(0)
    filename = f"Filtered_Reports_{username}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export_filtered_reports_excel", methods=["POST"])
def export_filtered_reports_excel():
    if "user" not in session:
        return jsonify({"error": "Not logged in"}), 401

    from datetime import datetime
    import pandas as pd
    from io import BytesIO
    from flask import send_file

    username = session["user"]
    user = User.query.filter_by(username=username).first()
    payload = request.get_json() or {}
    report_ids = payload.get("report_ids", [])

    if not report_ids:
        return jsonify({"error": "No report IDs provided"}), 400

    # Query only the reports visible in filtered view
    reports = db.session.query(Report, ContainerInfo).join(
        ContainerInfo, Report.container_id == ContainerInfo.id
    ).filter(Report.id.in_(report_ids)).all()

    if user.role != "admin":
        reports = [r for r in reports if r[0].username == username]

    if not reports:
        return jsonify({"error": "No matching reports found"}), 404

    data = []
    for r, c in reports:
        data.append({
            "Container No": c.container_no,
            "Line": c.line,
            "Size": c.size,
            "In Date": c.in_date,
            "Mfg Date": c.mfg_date,
            "GW": c.gw,
            "TW": c.tw,
            "CSC": c.csc,
            "Grand Total": getattr(r, "grand_total", ""),
            "File Type": r.file_type,
            "Date Generated": r.timestamp.strftime("%Y-%m-%d %H:%M")
        })

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Filtered Reports")
        ws = writer.sheets["Filtered Reports"]

        from openpyxl.styles import Alignment, Font
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    output.seek(0)
    filename = f"Filtered_Reports_{username}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------------- API Endpoints ----------------
@app.route("/api/descriptions")
def api_descriptions():
    category = request.args.get("category", "").strip()
    if not category:
        return jsonify([])

    rows = Tariff.query.filter_by(category=category).all()
    descriptions = {}
    for r in rows:
        desc = r.description
        if desc not in descriptions:
            descriptions[desc] = []
        if r.dimensions and r.dimensions.strip():
            descriptions[desc].append(r.dimensions)

    out = []
    for desc, dims in descriptions.items():
        out.append({
            "description": desc,
            "has_dimensions": len(dims) > 0,
            "dimensions": sorted(list(set(dims)))
        })
    return jsonify(out)

@app.route("/api/matcost")
def api_matcost():
    category = request.args.get("category", "").strip()
    description = request.args.get("description", "").strip()
    dimension = request.args.get("dimension", "").strip()

    query = Tariff.query.filter_by(category=category, description=description)

    # EVERY ADDITIONAL (special)
    if description.upper().startswith("EVERY ADDITIONAL"):
        row = query.first()
        if not row:
            return jsonify({"mat_cost": 0.0})
        try:
            user_val = float(dimension)
            return jsonify({"mat_cost": round(user_val * row.mat_cost, 2)})
        except:
            return jsonify({"mat_cost": 0.0})

    # If dimension provided, find exact match row
    if dimension:
        row = query.filter_by(dimensions=dimension).first()
        if row:
            return jsonify({"mat_cost": row.mat_cost})

    # fallback to first match
    row = query.first()
    return jsonify({"mat_cost": row.mat_cost if row else 0.0})

# ---------------- Exports (Excel + PDF) ----------------
@app.route("/export_excel", methods=["POST"])
def export_excel():
    payload = request.get_json()
    if payload is None:
        return jsonify({"error": "No data provided"}), 400

    entries = payload.get("entries") if isinstance(payload, dict) else payload
    container_info = payload.get("container") if isinstance(payload, dict) else {}
    if not container_info:
        container_info = session.get("estimation_data", {})
    
    # 🧹 STEP 1: Remove any unwanted "Grand Total" or "Totals" rows from frontend table
    cleaned_entries = []
    for e in entries:
        cat = str(e.get("category", "")).strip().lower()
        # Filter out frontend totals
        if cat not in ["total", "totals", "grand total"]:
            cleaned_entries.append(e)

    # 🧮 STEP 2: Convert to DataFrame safely
    if cleaned_entries:
        df = pd.DataFrame(cleaned_entries)
    else:
        df = pd.DataFrame(columns=["category", "description", "dimension", "mat_cost", "man_hrs", "lab_cost", "total"])

    username = session.get("user", "anonymous")
    container_no_raw = container_info.get("container_no") or "estimation"
    line = container_info.get("line", "N/A")
    container_no = safe_filename(container_no_raw)

    # ✅ Call the shared helper (where your existing Excel logic goes)
    file_path, grand_total = export_excel_internal(username, container_info, entries, line, container_no)

     # ✅ Log the export in DB
    report = Report(
        username=username,
        container_no=container_no,
        line=line,
        grand_total=grand_total,
        file_type="excel",
        file_path=file_path,
        container_id=session.get("container_id"),
        entries_json=entries
    )
    db.session.add(report)
    db.session.commit()

    # ✅ Return as downloadable file
    safe_line = safe_filename(line or "LINE")
    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{safe_line}_{container_no}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def export_excel_internal(username, container_info, entries, line, container_no):
    """
    Shared Excel generator used by both /export_excel and auto-regeneration.
    Returns (file_path, grand_total).
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    # ---- Normalize container_info for both dict and SQLAlchemy object ----
    if not isinstance(container_info, dict):
        container_info = {
            "container_no": getattr(container_info, "container_no", ""),
            "in_date": getattr(container_info, "in_date", ""),
            "mfg_date": getattr(container_info, "mfg_date", ""),
            "line": getattr(container_info, "line", ""),
            "size": getattr(container_info, "size", ""),
            "gw": getattr(container_info, "gw", ""),
            "tw": getattr(container_info, "tw", ""),
            "csc": getattr(container_info, "csc", "")
        }

    # ---- Normalize and prepare entries ----
    for i, row in enumerate(entries, start=1):
        row["sr"] = i
        row["mat_cost"] = round(float(row.get("mat_cost", 0) or 0), 2)
        row["lab_cost"] = round(float(row.get("lab_cost", 0) or 0), 2)
        row["total"] = round(float(row.get("total", 0) or 0), 2)

    df = pd.DataFrame(entries)

    # ✅ Remove any accidental "total" or "grand total" rows just in case
    df = df[~df['category'].astype(str).str.contains('total', case=False, na=False)]

    # Compute grand totals (for separate table)
    total_mat = df["mat_cost"].sum() if not df.empty else 0.0
    total_lab = df["lab_cost"].sum() if not df.empty else 0.0
    total_all = df["total"].sum() if not df.empty else 0.0


    # ---- Container Info Table ----
    container_headers = ["Container No", "In Date", "Mfg Date", "Line", "Size", "GW", "TW", "CSC"]
    container_data = [[
        container_info.get("container_no", ""),
        container_info.get("in_date", ""),
        container_info.get("mfg_date", ""),
        container_info.get("line", ""),
        container_info.get("size", ""),
        container_info.get("gw", ""),
        container_info.get("tw", ""),
        container_info.get("csc", "")
    ]]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet(title="Report")
        wb.active = ws

        # ---- Styles ----
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # ---- Header ----
        ws.merge_cells("A2:H2")
        ws["A2"] = "ALLIANCE MARINE TERMINAL"
        ws["A2"].font = Font(bold=True, size=16)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A3:H3")
        ws["A3"] = "At.Post-Jasai, Tel-Uran, Dist- Raigad Opp Indian Oil Petrol Pump, Navi Mumbai Pin-400 702"
        ws["A3"].font = Font(bold=True, size=11)
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        # ---- Container Info Table ----
        start_row = 5
        for col_num, header in enumerate(container_headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border

        for row_idx, row_data in enumerate(container_data, start=start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                if "container_no" in header.lower() or col_num == 1:  # assuming first column or name match
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # ---- Repair Table ----
        repair_start = start_row + len(container_data) + 3
        headers = ["SR", "Category", "Description", "Dimension", "MAT.COST", "MAN.HRS", "LAB.COST", "TOTAL"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=repair_start, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border

        for idx, row in enumerate(df.itertuples(index=False), start=repair_start + 1):
            ws.cell(row=idx, column=1, value=getattr(row, "sr", ""))
            ws.cell(row=idx, column=2, value=getattr(row, "category", ""))
            ws.cell(row=idx, column=3, value=getattr(row, "description", ""))
            ws.cell(row=idx, column=4, value=getattr(row, "dimension", ""))
            ws.cell(row=idx, column=5, value=getattr(row, "mat_cost", ""))
            ws.cell(row=idx, column=6, value=getattr(row, "man_hrs", ""))
            ws.cell(row=idx, column=7, value=getattr(row, "lab_cost", ""))
            ws.cell(row=idx, column=8, value=getattr(row, "total", ""))

            for c in range(1, 9):
                cell = ws.cell(row=idx, column=c)
                if not isinstance(cell, MergedCell):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

       # ---- Grand Total Table ----
        grand_row = repair_start + len(df) + 2

        ws["B" + str(grand_row)] = "GRAND TOTAL"
        ws["B" + str(grand_row)].font = Font(bold=True)
        ws["B" + str(grand_row)].alignment = Alignment(horizontal="right", vertical="center")

        # Calculate totals
        mat_total = df["mat_cost"].sum() if "mat_cost" in df.columns else 0
        lab_total = df["lab_cost"].sum() if "lab_cost" in df.columns else 0
        grand_total = df["total"].sum() if "total" in df.columns else mat_total + lab_total

        ws[f"E{grand_row}"] = mat_total
        ws[f"G{grand_row}"] = lab_total
        ws[f"H{grand_row}"] = grand_total

        for col in ["E", "G", "H"]:
            ws[f"{col}{grand_row}"].font = Font(bold=True)
            ws[f"{col}{grand_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Apply borders to grand total table row
        for col_idx in range(1, 9):
            ws.cell(row=grand_row, column=col_idx).border = border

        # ---- Approved Amount + Date Generated (combined 2-row table) ----
        approved_row = grand_row + 3
        date_row = approved_row + 1

        # Row 1: Approved Amount
        ws[f"A{approved_row}"] = "Approved Amount"
        ws[f"A{approved_row}"].font = Font(bold=True)
        ws[f"A{approved_row}"].alignment = Alignment(horizontal="right", vertical="center")

        ws[f"B{approved_row}"] = ""  # Leave empty for manual input
        ws[f"B{approved_row}"].font = Font(bold=True)
        ws[f"B{approved_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: Date Generated
        ws[f"A{date_row}"] = "Date Generated"
        ws[f"A{date_row}"].font = Font(bold=True)
        ws[f"A{date_row}"].alignment = Alignment(horizontal="right", vertical="center")

        ws[f"B{date_row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f"B{date_row}"].font = Font(bold=True)
        ws[f"B{date_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Apply borders to both rows
        for r in [approved_row, date_row]:
            ws[f"A{r}"].border = border
            ws[f"B{r}"].border = border

        # ---- Auto-fit ----
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    vlen = len(str(cell.value))
                    if vlen > max_length:
                        max_length = vlen
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[col_letter].width = max(10, min(max_length + 4, 50))

    # ---- Save file ----
    output.seek(0)
    user_dir = ensure_user_dir(username)
    file_path = os.path.join(user_dir, f"{line}_{container_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    with open(file_path, "wb") as f:
        f.write(output.getvalue())

    return file_path, float(total_all)
           
@app.route("/export_pdf", methods=["POST"])
def export_pdf():
    payload = request.get_json()
    if payload is None:
        return jsonify({"error": "No data provided"}), 400

    entries = payload.get("entries") if isinstance(payload, dict) else payload
    container_info = payload.get("container") if isinstance(payload, dict) else {}
    if not container_info:
        container_info = session.get("estimation_data", {})

    username = session.get("user", "anonymous")
    container_no_raw = container_info.get("container_no") or "estimation"
    line = container_info.get("line", "N/A")
    container_no = safe_filename(container_no_raw)

    # ✅ Call shared helper (which contains your existing logic)
    file_path, grand_total = export_pdf_internal(username, container_info, entries, line, container_no)

 # ✅ Save to DB
    report = Report(
        username=username,
        container_no=container_no,
        line=line,
        grand_total=grand_total,
        file_type="pdf",
        file_path=file_path,
        container_id=session.get("container_id"),
        entries_json=entries  # 🆕 store repair table entries
    )
    db.session.add(report)
    db.session.commit()

    # ✅ Return file
    safe_line = safe_filename(line or "LINE")
    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"{safe_line}_{container_no}.pdf",
        mimetype="application/pdf",
    )

def export_pdf_internal(username, container_info, entries, line, container_no):

    if not isinstance(container_info, dict):
        container_info = {
            "container_no": getattr(container_info, "container_no", ""),
            "in_date": getattr(container_info, "in_date", ""),
            "mfg_date": getattr(container_info, "mfg_date", ""),
            "line": getattr(container_info, "line", ""),
            "size": getattr(container_info, "size", ""),
            "gw": getattr(container_info, "gw", ""),
            "tw": getattr(container_info, "tw", ""),
            "csc": getattr(container_info, "csc", "")
        }
        
    styles = getSampleStyleSheet()
    centered = ParagraphStyle(name="centered", parent=styles['Normal'], alignment=TA_CENTER)
    right = ParagraphStyle(name="right", parent=styles["Normal"], alignment=TA_RIGHT)

    # Container info table
    info_headers = ["Container No", "In Date", "Mfg Date", "Line", "Size", "GW", "TW", "CSC"]
    info_values = []
    for key in ["container_no", "in_date", "mfg_date", "line", "size", "gw", "tw", "csc"]:
        val = container_info.get(key, "")
        if key == "container_no":
            info_values.append(Paragraph(f"<b>{val}</b>", centered))
        else:
            info_values.append(Paragraph(str(val), centered))
    info_table = [info_headers, info_values]

    # Estimations table
    headers = ["Category", "Description", "Dimension", "MAT.COST", "MAN.HRS", "LAB.COST", "TOTAL"]
    table_data = [headers]
    total_sum = total_mat = total_lab = 0.0
    for row in entries:
        mat_cost = float(row.get("mat_cost", 0) or 0)
        lab_cost = float(row.get("lab_cost", 0) or 0)
        man_hrs = row.get("man_hrs", "")
        total = mat_cost + lab_cost
        total_mat += mat_cost
        total_lab += lab_cost
        total_sum += total
        table_data.append([
            Paragraph(row.get("category", ""), styles['Normal']),
            Paragraph(row.get("description", ""), styles['Normal']),
            Paragraph(str(row.get("dimension", "")), styles['Normal']),
            f"{mat_cost:.2f}",
            str(man_hrs),
            f"{lab_cost:.2f}",
            f"{total:.2f}"
        ])

    table_data.append([
        Paragraph("<b>GRAND TOTAL</b>", styles['Normal']), "", "",
        f"{total_mat:.2f}", "",
        f"{total_lab:.2f}",
        f"{total_sum:.2f}"
    ])

    # Build PDF
    buffer = io.BytesIO()
    page_w, page_h = landscape(A4)
    leftMargin, rightMargin, topMargin, bottomMargin = 40, 40, 80, 60
    usable_width = page_w - leftMargin - rightMargin

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=leftMargin, rightMargin=rightMargin,
                            topMargin=topMargin, bottomMargin=bottomMargin)
    elems = []

    elems.append(Paragraph("<para align='center'><b>ALLIANCE MARINE TERMINAL</b></para>", styles['Title']))
    elems.append(Paragraph(
        "<para align='center'><b>At.Post-Jasai, Tel-Uran, Dist- Raigad Opp Indian Oil Petrol Pump, Navi Mumbai Pin-400 702</b></para>",
        styles['Heading4']
    ))
    elems.append(Spacer(1, 30))

    # Container info table
    colw_info = [usable_width / len(info_headers)] * len(info_headers)
    info_tbl = Table(info_table, colWidths=colw_info, hAlign='CENTER')
    info_tbl.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f2f2f2")),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elems.append(info_tbl)
    elems.append(Spacer(1, 24))

    # Estimations table
    colw_est = [
        usable_width * 0.12, usable_width * 0.38, usable_width * 0.10,
        usable_width * 0.10, usable_width * 0.08, usable_width * 0.10,
        usable_width * 0.12
    ]
    tbl = Table(table_data, colWidths=colw_est, repeatRows=1, hAlign='CENTER')
    tbl.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#d9e6f6")),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#f2f2f2")),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 24))

    # ---- Approved Amount + Date Generated ----
    approved_data = [
        ["Approved Amount", ""],
        ["Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    approved_tbl = Table(approved_data, colWidths=[200, 250], hAlign="CENTER")
    approved_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems.append(approved_tbl)

    doc.build(elems)

    # ---------------- Save + log ----------------
    user_dir = ensure_user_dir(username)
    file_path = os.path.join(
        user_dir,
        f"{line}_{container_no}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    # Save the PDF file
    with open(file_path, "wb") as f:
        f.write(buffer.getvalue())

    # ✅ Link report to the container and include grand total
    container_id = session.get("container_id")

    # safely calculate or fallback
    grand_total = 0.0
    if "total_all" in locals():
        try:
            grand_total = float(total)
        except Exception:
            grand_total = 0.0
    elif "entries" in locals() and entries:
        for e in entries:
            try:
                grand_total += float(e.get("mat_cost", 0)) + float(e.get("lab_cost", 0))
            except Exception:
                continue

    return file_path, float(total_sum)            

def regenerate_pdf(report, container, entries):
    """Auto-regenerate missing PDF using the existing export logic."""
    print(f"[Auto-regen] Rebuilding PDF for {report.container_no}...")
    return export_pdf_internal(report.username, container, entries, container.line, container.container_no)


def regenerate_excel(report, container, entries):
    """Auto-regenerate missing Excel using the existing export logic."""
    print(f"[Auto-regen] Rebuilding Excel for {report.container_no}...")
    return export_excel_internal(report.username, container, entries, container.line, container.container_no)

def get_entries_for_report(container_no):
    """
    Fetch all repair entries (tariff-based + manual) for a given container.
    Adjust this query based on how you currently store entries.
    """
    entries = []

    # Example: if you store repair entries as a JSON field or separate table
    # you can query it here. For now, we'll assume you derive from tariffs.
    tariffs = Tariff.query.all()
    for t in tariffs:
        entries.append({
            "category": t.category,
            "description": t.description,
            "dimension": t.dimensions or "",
            "mat_cost": float(t.mat_cost or 0),
            "man_hrs": 0,
            "lab_cost": 0,
            "total": float(t.mat_cost or 0)
        })
    return entries

# ---------- New helper APIs ----------

@app.route("/api/lines")
def api_lines():
    """Return distinct client lines for dropdown."""
    lines = [l[0] for l in db.session.query(ContainerInfo.line).distinct().all() if l[0]]
    lines.sort()
    return jsonify(lines)


@app.route("/api/container_info")
def api_container_info():
    """Return stored container details by container_no."""
    container_no = request.args.get("container_no", "").strip()
    if not container_no:
        return jsonify({"error": "Missing container_no"}), 400

    record = ContainerInfo.query.filter_by(container_no=container_no).order_by(ContainerInfo.id.desc()).first()
    if not record:
        return jsonify({"found": False})

    data = {
        "found": True,
        "line": record.line,
        "mfg_date": record.mfg_date,
        "size": record.size,
        "gw": record.gw,
        "tw": record.tw,
        "csc": record.csc,
    }
    return jsonify(data)

# --- Admin panel setup ---
from admin import init_admin

init_admin(app, db, User, Tariff, ContainerInfo, Report)
print("✅ Flask-Admin successfully initialized at /admin_panel")

# ---------------- Run ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)


